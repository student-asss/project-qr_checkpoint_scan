# server.py
from flask import Flask, request, jsonify, redirect, url_for
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
import socket
from threading import Lock

app = Flask(__name__)

# Configuration
DATA_DIR = os.path.join(os.path.expanduser('~'), 'Desktop', 'Checkpoint - Finished product')
current_date =  datetime.now().strftime('%Y%m%d')
EXCEL_FILE = os.path.join(DATA_DIR, f'./data/qr_checkpoints_{current_date}.xlsx')
PORT = 9500  # Changed to a more reliable port

# Ensure data directory exists
os.makedirs(DATA_DIR, exist_ok=True)

# Thread lock for safe Excel operations
excel_lock = Lock()

def get_local_ip():
    """Get local IP without external dependencies"""
    try:
        # Method 1: UDP connection (works on most systems)
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as s:
            s.connect(('8.8.8.8', 80))
            ip = s.getsockname()[0]
            if not ip.startswith('127.'):
                return ip
    except:
        pass
    
    try:
        # Method 2: Hostname resolution
        hostname = socket.gethostname()
        ip = socket.gethostbyname(hostname)
        if not ip.startswith('127.'):
            return ip
    except:
        pass
    
    # Fallback to localhost
    return '127.0.0.1'

def initialize_excel():
    with excel_lock:
        """Create Excel file with headers if it doesn't exist"""
        if not os.path.exists(EXCEL_FILE):
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Checkpoints"
                ws.append(["QR Code", "Scan Date", "Scan Time"])
                wb.save(EXCEL_FILE)
                print(f"Created new Excel file at: {EXCEL_FILE}")
            except Exception as e:
                print(f"Error creating Excel file: {e}")

def record_scan(qr_code):
    """Record a scan in the Excel file"""
    now = datetime.now()
    scan_date = now.strftime("%Y-%m-%d")
    scan_time = now.strftime("%H:%M:%S")
    timestamp = now.isoformat()

    with excel_lock:
        try:
            print(f"Attempting to record scan: {qr_code}")
            if os.path.exists(EXCEL_FILE):
                wb = load_workbook(EXCEL_FILE)
            else:
                wb = Workbook()
            
            ws = wb.active
            ws.append([qr_code, scan_date, scan_time])
            for column in ['A', 'B', 'C']:
                ws.column_dimensions[column].width = 20
            wb.save(EXCEL_FILE)
            print("Scan recorded successfully")
            return True
        except Exception as e:
            print(f"Error saving scan: {e}")
            return False

@app.route('/scan', methods=['POST', 'GET'])
def handle_scan():
    """Handle scan submissions and show success page"""
    if request.method == 'POST':
        qr_code = request.form.get('qr_code') or request.json.get('qr_code')
    else:  # GET
        qr_code = request.args.get('qr_code')

    if not qr_code:
        return jsonify({"error": "QR code data is required"}), 400

    success = record_scan(qr_code)
    server_ip = get_local_ip()
    
    if success:
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Scan Successful</title>
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <style>
                body {{ font-family: Arial, sans-serif; max-width: 800px; margin: 0 auto; padding: 20px; text-align: center; }}
                .success-message {{ 
                    background-color: #dff0d8;
                    color: #3c763d;
                    padding: 20px;
                    border-radius: 5px;
                    margin: 20px 0;
                }}
                .scan-info {{
                    margin: 20px 0;
                    padding: 15px;
                    background-color: #f9f9f9;
                    border-radius: 5px;
                }}
                .btn {{
                    display: inline-block;
                    padding: 10px 20px;
                    background-color: #337ab7;
                    color: white;
                    text-decoration: none;
                    border-radius: 5px;
                    margin-top: 20px;
                }}
            </style>
        </head>
        <body>
            <h1>Sunset Mkuu Checkpoint System</h1>
            <div class="success-message">
                <h2>✓ Scan Recorded Successfully!</h2>
            </div>
            <div class="scan-info">
                <p><strong>Checkpoint ID:</strong> {qr_code}</p>
                <p><strong>Time:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            </div>
            <a href="http://{server_ip}:{PORT}" class="btn">Return to Scanner</a>
        </body>
        </html>
        """
        return html_content
    else:
        return jsonify({"error": "Failed to record scan"}), 500

@app.route('/')
def index():
    """Serve the monitoring page"""
    server_ip = get_local_ip()
    network_url = f"http://{server_ip}:{PORT}"
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Sunset Mkuu</title>
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <style>
            body {{ font-family: Arial, sans-serif; max-width: 800px; margin: 0 auto; padding: 20px; }}
            .instructions {{ background-color: #f0f8ff; padding: 15px; border-radius: 5px; margin-bottom: 20px; }}
            .url-box {{ background: #e0e0e0; padding: 10px; border-radius: 5px; word-break: break-all; }}
            .recent-scans {{ margin-top: 30px; }}
            table {{ width: 100%; border-collapse: collapse; }}
            th, td {{ padding: 8px; text-align: left; border-bottom: 1px solid #ddd; }}
        </style>
    </head>
    <body>
        <h1>Sunset Mkuu - QR Checkpoint System</h1>
        <div class="instructions">
            <h3>How to use:</h3>
            <p>1. Connect your phone to the same WiFi network as this computer</p>
            <p>2. Open any QR scanner app on your phone</p>
            <p>3. Scan a QR code containing:</p>
            <div class="url-box">{network_url}/scan?qr_code=YOUR_DATA</div>
            <p>4. The scan will be recorded in the Excel file</p>
            <p><strong>Server running on:</strong></p>
            <div class="url-box">{network_url}</div>
        </div>
    </body>
    </html>
    """
    return html_content

if __name__ == '__main__':
    initialize_excel()
    local_ip = get_local_ip()
    
    print("\n" + "="*50)
    print(f"Local access: http://localhost:{PORT}")
    print(f"Network access: http://{local_ip}:{PORT}")
    print("="*50 + "\n")
    
    print("Troubleshooting Tips:")
    print("1. Ensure both devices are on the same WiFi network")
    print("2. Try accessing the network URL from your phone's browser first")
    print("3. If connection fails, check your firewall/antivirus settings")
    print("4. Make sure no other application is using port", PORT)
    
    try:
        app.run(host='0.0.0.0', port=PORT, threaded=True, debug=True)
    except Exception as e:
        print(f"\nFailed to start server: {e}")
        print("Try these solutions:")
        print("1. Use a different port (change PORT variable)")
        print("2. Run as administrator")
        print("3. Check for conflicting applications")